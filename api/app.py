import copy
import io
import os
import secrets
import smtplib
import ssl
import string
import threading
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

import bcrypt
import jwt
from flask import Flask, jsonify, request, send_file, send_from_directory, g, redirect
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from collector.runner import run_collect, get_status
from collector.expiry import run_expiry_collect, get_expiry_status  # kept for /api/expiry/refresh endpoint
from providers import create_provider
from storage.mongo import MongoStore

app = Flask(__name__)
_cors_origins = [o.strip() for o in os.environ.get("CORS_ORIGINS", "").split(",") if o.strip()]
CORS(app, origins=_cors_origins if _cors_origins else ["http://localhost:5173", "http://localhost:5002"])

_cfg: dict = {}
_mongo_cfg: dict = {}

_SECRET_KEYS = {"jwt_secret", "password", "sk"}


def _hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()


def _check_password(plain: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(plain.encode(), hashed.encode())
    except Exception:
        return False


def _gen_random_password(length: int = 12) -> str:
    chars = string.ascii_letters + string.digits + "!@#$%^&*"
    return "".join(secrets.choice(chars) for _ in range(length))


def init_app(config: dict, mongo_cfg: dict):
    global _cfg, _mongo_cfg
    _cfg = config
    _mongo_cfg = mongo_cfg
    # 首次启动初始化：admin 用户 + 预设权限组
    try:
        store = _store()
        admin_pw = config.get("auth", {}).get("password", "Admin@123")
        store.init_admin_user(config.get("auth", {}).get("username", "admin"),
                              _hash_password(admin_pw))
        store.init_preset_groups()
    except Exception as e:
        print(f"[init] 用户/权限组初始化失败: {e}")
    return app


def _store() -> MongoStore:
    return MongoStore(_mongo_cfg["host"], _mongo_cfg["port"], _mongo_cfg["database"])


def _require_auth(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        token = request.headers.get("Authorization", "").removeprefix("Bearer ").strip()
        if not token:
            token = request.args.get("_t", "").strip()
        if not token:
            return jsonify({"code": 401, "msg": "未登录"}), 401
        try:
            payload = jwt.decode(token, _cfg["auth"]["jwt_secret"], algorithms=["HS256"])
            g.username = payload["sub"]
            g.role = payload.get("role", "user")
            g.menus = payload.get("menus", [])
            g.actions = payload.get("actions", [])
        except jwt.ExpiredSignatureError:
            return jsonify({"code": 401, "msg": "登录已过期"}), 401
        except Exception:
            return jsonify({"code": 401, "msg": "Token无效"}), 401
        return f(*args, **kwargs)
    return wrapper


def _require_admin(f):
    @wraps(f)
    @_require_auth
    def wrapper(*args, **kwargs):
        if g.role != "admin":
            return jsonify({"code": 403, "msg": "权限不足，需要管理员"}), 403
        return f(*args, **kwargs)
    return wrapper


def _require_action(action: str):
    def decorator(f):
        @wraps(f)
        @_require_auth
        def wrapper(*args, **kwargs):
            if g.role == "admin" or "*" in g.actions or action in g.actions:
                return f(*args, **kwargs)
            return jsonify({"code": 403, "msg": f"权限不足，需要操作权限: {action}"}), 403
        return wrapper
    return decorator


# ── 认证 ──────────────────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def login():
    data = request.json or {}
    auth_cfg = _cfg["auth"]
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
    ua = request.headers.get("User-Agent", "")
    username = data.get("username", "")
    password = data.get("password", "")

    store = _store()
    user = store.get_user_by_username(username)

    def _fail(reason="用户名或密码错误"):
        try:
            store.record_login(username or "unknown", ip, ua, success=False, fail_reason=reason)
        except Exception:
            pass
        return jsonify({"code": 401, "msg": reason}), 401

    if not user:
        return _fail()
    if not user.get("is_active", True):
        return _fail("账号已被禁用")
    if not _check_password(password, user.get("password_hash", "")):
        return _fail()

    role = user.get("role", "user")
    if role == "admin":
        menus = ["*"]
        actions = ["*"]
    else:
        group = store.get_permission_group(user.get("group_id") or "") or {}
        menus = group.get("menus", [])
        actions = group.get("actions", [])

    expire = int(time.time()) + auth_cfg.get("token_expire_hours", 24) * 3600
    payload = {
        "sub": username,
        "role": role,
        "menus": menus,
        "actions": actions,
        "exp": expire,
    }
    token = jwt.encode(payload, auth_cfg["jwt_secret"], algorithm="HS256")

    try:
        store.record_login(username, ip, ua, success=True)
        store.update_last_login(username)
    except Exception:
        pass

    return jsonify({"code": 0, "token": token, "username": username, "role": role,
                    "menus": menus, "actions": actions, "v": 2})


@app.route("/api/auth/verify-password", methods=["POST"])
@_require_auth
def verify_password():
    data = request.json or {}
    user = _store().get_user_by_username(g.username)
    if not user or not _check_password(data.get("password", ""), user.get("password_hash", "")):
        return jsonify({"code": 403, "msg": "密码错误"}), 403
    return jsonify({"code": 0})


@app.route("/api/login-history")
@_require_auth
def login_history():
    history = _store().get_login_history(g.username, limit=20)
    return jsonify({"code": 0, "data": history})


@app.route("/api/version")
def version():
    try:
        ver_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), "version.txt")
        with open(ver_file) as f:
            ver = f.read().strip()
    except Exception:
        ver = "dev"
    return jsonify({"code": 0, "version": ver})


# ── 筛选选项 ──────────────────────────────────────────────────────────────────

@app.route("/api/filter-options")
@_require_auth
def filter_options():
    opts = _store().get_filter_options()
    return jsonify({"code": 0, "data": opts})


# ── 大屏汇总 ──────────────────────────────────────────────────────────────────

@app.route("/api/summary")
@_require_auth
def summary():
    days = int(request.args.get("days", 1))
    data = _store().get_summary(days)
    return jsonify({"code": 0, "data": data})


# ── 详情列表 ──────────────────────────────────────────────────────────────────

@app.route("/api/metrics")
@_require_auth
def metrics():
    days = int(request.args.get("days", 7))
    filters = {
        "cloud": request.args.get("cloud", ""),
        "provider": request.args.get("provider", ""),
        "region": request.args.get("region", ""),
        "project": request.args.get("project", ""),
        "group": request.args.get("group", ""),
        "name": request.args.get("name", ""),
        "ip": request.args.get("ip", ""),
    }
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))
    sort_field = request.args.get("sort_field", "cpu_avg")
    sort_order = request.args.get("sort_order", "desc")

    all_data = _store().query_metrics(days, filters, sort_field, sort_order)
    total = len(all_data)
    paged = all_data[(page - 1) * page_size: page * page_size]
    return jsonify({"code": 0, "data": paged, "total": total, "page": page, "page_size": page_size})


# ── 导出 Excel ────────────────────────────────────────────────────────────────

@app.route("/api/export")
@_require_auth
def export_excel():
    days = int(request.args.get("days", 7))
    filters = {
        "cloud": request.args.get("cloud", ""),
        "provider": request.args.get("provider", ""),
        "region": request.args.get("region", ""),
        "project": request.args.get("project", ""),
        "group": request.args.get("group", ""),
    }
    data = _store().query_metrics(days, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "云监控数据"

    headers = ["序号", "名称", "云账号", "云厂商", "区域", "项目", "部门", "二级部门", "负责人",
               "内网IP", "外网IP", "CPU核数", "内存(GB)", "CPU均值(%)", "CPU峰值(%)",
               "内存均值(%)", "内存峰值(%)", "磁盘最大(%)", "磁盘详情", "出带宽均值(Mbps)", "入带宽均值(Mbps)"]

    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    high_fill = PatternFill(start_color="FF7043", end_color="FF7043", fill_type="solid")
    warn_fill = PatternFill(start_color="FFA726", end_color="FFA726", fill_type="solid")

    for idx, row in enumerate(data, 1):
        disk_detail_str = " | ".join(
            f"{d['device']}:{d['usage']}%" for d in (row.get("disk_details") or [])
        ) or ""
        values = [
            idx, row.get("name"), row.get("cloud"), row.get("provider"),
            row.get("region"), row.get("project"), row.get("group"),
            row.get("last_group"), row.get("manager"),
            row.get("n_ip"), row.get("w_ip"),
            row.get("cpus"), row.get("ram"),
            row.get("cpu_avg"), row.get("cpu_max"),
            row.get("mem_avg"), row.get("mem_max"),
            row.get("disk_avg"), disk_detail_str,
            row.get("out_avg"), row.get("in_avg"),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            if col in (14, 16) and isinstance(val, (int, float)):
                if val >= 80:
                    cell.fill = high_fill
                elif val >= 60:
                    cell.fill = warn_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"云监控数据_{time.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 到期资源 ──────────────────────────────────────────────────────────────────

@app.route("/api/expiry")
@_require_auth
def expiry_list():
    days = int(request.args.get("days", 30))
    filters = {
        "cloud": request.args.get("cloud", ""),
        "provider": request.args.get("provider", ""),
        "resource_type": request.args.get("resource_type", ""),
        "group": request.args.get("group", ""),
        "project": request.args.get("project", ""),
        "auto_renew": request.args.get("auto_renew", ""),
    }
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))
    sort_field = request.args.get("sort_field", "expire_time")
    sort_order = request.args.get("sort_order", "asc")
    all_data = _store().query_expiry_resources(days, filters, sort_field, sort_order)
    total = len(all_data)
    paged = all_data[(page - 1) * page_size: page * page_size]
    return jsonify({"code": 0, "data": paged, "total": total, "page": page, "page_size": page_size})


@app.route("/api/expiry/filter-options")
@_require_auth
def expiry_filter_options():
    opts = _store().get_expiry_filter_options()
    return jsonify({"code": 0, "data": opts})


@app.route("/api/expiry/refresh", methods=["POST"])
@_require_action("refresh_expiry")
def expiry_refresh():
    cloud_configs = _cfg.get("clouds", [])
    mongo_cfg = _cfg["mongodb"]
    t = threading.Thread(
        target=run_expiry_collect,
        args=(cloud_configs, mongo_cfg, 30),
        daemon=True,
    )
    t.start()
    return jsonify({"code": 0, "msg": "到期资源采集已启动"})


@app.route("/api/expiry/status")
@_require_auth
def expiry_status():
    return jsonify({"code": 0, "data": get_expiry_status()})


@app.route("/api/expiry/export")
@_require_auth
def expiry_export():
    days = int(request.args.get("days", 30))
    filters = {
        "cloud": request.args.get("cloud", ""),
        "provider": request.args.get("provider", ""),
        "resource_type": request.args.get("resource_type", ""),
        "group": request.args.get("group", ""),
        "project": request.args.get("project", ""),
        "auto_renew": request.args.get("auto_renew", ""),
    }
    data = _store().query_expiry_resources(days, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "资源到期列表"
    headers = ["序号", "资源名称", "资源类型", "云账号", "云厂商", "区域", "项目",
               "部门", "二级部门", "负责人", "计费方式", "状态", "自动续费", "到期时间", "剩余天数"]
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    red_fill = PatternFill(start_color="FF5252", end_color="FF5252", fill_type="solid")
    orange_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    yellow_fill = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

    for idx, row in enumerate(data, 1):
        d_left = row.get("days_left", 99)
        values = [
            idx, row.get("name"), row.get("resource_type"),
            row.get("cloud"), row.get("provider"), row.get("region"), row.get("project"),
            row.get("group"), row.get("last_group"), row.get("manager"),
            row.get("charging_mode"), row.get("status"), row.get("auto_renew", "未知"),
            row.get("expire_time"), d_left,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            if col == 15:
                if d_left <= 7:
                    cell.fill = red_fill
                elif d_left <= 15:
                    cell.fill = orange_fill
                elif d_left <= 30:
                    cell.fill = yellow_fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"资源到期列表_{time.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── 采集控制 ──────────────────────────────────────────────────────────────────

@app.route("/api/collect", methods=["POST"])
@_require_action("collect")
def trigger_collect():
    cloud_configs = _cfg.get("clouds", [])
    mongo_cfg = _cfg["mongodb"]
    hours = _cfg.get("schedule", {}).get("hours", 24)

    threading.Thread(target=run_collect, args=(cloud_configs, mongo_cfg, hours), daemon=True).start()
    return jsonify({"code": 0, "msg": "采集任务已启动"})


@app.route("/api/collect/status")
@_require_auth
def collect_status():
    return jsonify({"code": 0, "data": get_status()})


# ── 刷新项目信息（从 DB 重新 enrich 所有实例）────────────────────────────────

_refresh_projects_status: dict = {"status": "idle", "result": None, "error": None}


def _run_refresh_projects(mongo_cfg: dict):
    global _refresh_projects_status
    _refresh_projects_status = {"status": "running", "result": None, "error": None}
    try:
        store = MongoStore(mongo_cfg["host"], mongo_cfg["port"], mongo_cfg["database"])
        project_map = store.get_project_map_from_db()
        result = store.refresh_project_info(project_map)
        _refresh_projects_status = {"status": "done", "result": result, "error": None}
    except Exception as e:
        print(f"[refresh_projects] 刷新失败: {e}")
        _refresh_projects_status = {"status": "error", "result": None, "error": "刷新失败，请查看服务日志"}


@app.route("/api/refresh-projects", methods=["POST"])
@_require_action("refresh_projects")
def refresh_projects():
    if _refresh_projects_status.get("status") == "running":
        return jsonify({"code": 0, "msg": "刷新任务已在运行中"})
    threading.Thread(target=_run_refresh_projects, args=(_mongo_cfg,), daemon=True).start()
    return jsonify({"code": 0, "msg": "刷新任务已启动"})


@app.route("/api/refresh-projects/status")
@_require_auth
def refresh_projects_status():
    return jsonify({"code": 0, "data": _refresh_projects_status})


# ── 项目维护 CRUD ─────────────────────────────────────────────────────────────

@app.route("/api/projects")
@_require_auth
def list_projects():
    return jsonify({"code": 0, "data": _store().get_all_projects()})


@app.route("/api/projects", methods=["POST"])
@_require_auth
def create_project():
    data = request.json or {}
    if not data.get("project"):
        return jsonify({"code": 400, "msg": "项目名不能为空"}), 400
    pid = _store().upsert_project(data)
    return jsonify({"code": 0, "data": {"_id": pid}})


@app.route("/api/projects/<project_id>", methods=["PUT"])
@_require_auth
def update_project(project_id):
    data = request.json or {}
    data["_id"] = project_id
    _store().upsert_project(data)
    return jsonify({"code": 0})


@app.route("/api/projects/<project_id>", methods=["DELETE"])
@_require_auth
def delete_project(project_id):
    ok = _store().delete_project(project_id)
    return jsonify({"code": 0 if ok else 404})


# ── 账户余额 ──────────────────────────────────────────────────────────────────

@app.route("/api/balance")
@_require_auth
def account_balance():
    cloud_configs = _cfg.get("clouds", [])
    data = []
    for cc in cloud_configs:
        provider = create_provider(cc)
        if not provider:
            continue
        balance = provider.get_balance()
        if balance is not None:
            data.append({
                "cloud": cc.get("name", ""),
                "provider": cc.get("provider", ""),
                "cash_balance": balance.cash_balance,
                "coupon_amount": balance.coupon_amount,
                "credit_limit": balance.credit_limit,
                "owe_amount": balance.owe_amount,
                "available_amount": round(balance.available_amount, 2),
                "currency": balance.currency,
            })
    return jsonify({"code": 0, "data": data})


# ── 系统配置 ──────────────────────────────────────────────────────────────────

@app.route("/api/config")
@_require_auth
def get_config():
    data = _cfg_sanitized()
    # MongoDB 连接来自环境变量，只读展示
    data["mongodb"] = {**_mongo_cfg, "_readonly": True}
    return jsonify({"code": 0, "data": data})


@app.route("/api/config", methods=["POST"])
@_require_admin
def save_config():
    global _cfg
    raw = request.json or {}
    if not raw:
        return jsonify({"code": 400, "msg": "空配置"}), 400

    new_data = _strip_placeholders(raw)
    new_data.pop("mongodb", None)  # MongoDB 连接不可通过 API 修改

    updated = copy.deepcopy(_cfg)
    _deep_merge(updated, new_data)

    # 修复 clouds.sk：按 AK 匹配还原被脱敏的 SK
    if "clouds" in raw:
        orig_sk_by_ak = {c.get("ak", ""): c.get("sk", "") for c in _cfg.get("clouds", [])}
        for cloud in updated.get("clouds", []):
            if not cloud.get("sk") and cloud.get("ak") in orig_sk_by_ak:
                cloud["sk"] = orig_sk_by_ak[cloud["ak"]]

    restart_needed = _check_restart_needed(_cfg, updated)

    try:
        _store().save_app_config(updated)
    except Exception as e:
        return jsonify({"code": 500, "msg": f"保存配置失败: {e}"}), 500

    _cfg = updated
    return jsonify({"code": 0, "msg": "配置已保存", "restart_needed": restart_needed})


def _cfg_sanitized() -> dict:
    """返回配置，敏感字段（密码/SK/JWT密钥）全部脱敏为 ****"""
    def _mask(obj):
        if isinstance(obj, dict):
            return {k: "****" if k in _SECRET_KEYS else _mask(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_mask(i) for i in obj]
        return obj
    cfg = copy.deepcopy(_cfg)
    cfg.pop("project_excel", None)
    return _mask(cfg)


def _get_cfg_value_by_path(path: str):
    """通过点号分隔路径安全读取配置值（支持列表下标）"""
    parts = path.split(".")
    cur = _cfg
    for part in parts:
        if isinstance(cur, dict):
            cur = cur.get(part)
        elif isinstance(cur, list):
            try:
                cur = cur[int(part)]
            except (ValueError, IndexError):
                return None
        else:
            return None
        if cur is None:
            return None
    return cur


def _strip_placeholders(obj):
    """递归移除 **** 占位符，防止前端未修改的脱敏字段覆盖实际配置"""
    if isinstance(obj, dict):
        return {k: _strip_placeholders(v) for k, v in obj.items() if v != "****"}
    if isinstance(obj, list):
        return [_strip_placeholders(i) for i in obj]
    return obj


def _deep_merge(base: dict, override: dict):
    for k, v in override.items():
        if k in base and isinstance(base[k], dict) and isinstance(v, dict):
            _deep_merge(base[k], v)
        else:
            base[k] = v


def _check_restart_needed(old: dict, new: dict) -> list[str]:
    reasons = []
    if old.get("server", {}).get("port") != new.get("server", {}).get("port"):
        reasons.append("服务端口")
    if old.get("schedule", {}).get("time") != new.get("schedule", {}).get("time"):
        reasons.append("定时采集时间")
    return reasons


# ── 审计 ─────────────────────────────────────────────────────────────────────

_ALLOWED_SECRET_PATHS = {
    "auth.password",
    "auth.jwt_secret",
    "email.password",
}


@app.route("/api/audit/verify-secret", methods=["POST"])
@_require_auth
def verify_secret():
    data = request.json or {}
    password = data.get("password", "")
    config_path = data.get("config_path", "")

    user = _store().get_user_by_username(g.username)
    if not user or not _check_password(password, user.get("password_hash", "")):
        return jsonify({"code": 403, "msg": "密码错误"}), 403

    # 允许 auth.password / auth.jwt_secret / email.password / clouds.N.sk
    is_allowed = config_path in _ALLOWED_SECRET_PATHS or (
        config_path.startswith("clouds.") and config_path.endswith(".sk")
    )
    if not is_allowed:
        return jsonify({"code": 400, "msg": "不支持查看此配置项"}), 400

    value = _get_cfg_value_by_path(config_path)
    if value is None:
        return jsonify({"code": 404, "msg": "配置项不存在"}), 404

    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
    try:
        _store().record_secret_view(g.username, ip, config_path)
    except Exception:
        pass

    return jsonify({"code": 0, "value": str(value)})


@app.route("/api/audit/logs")
@_require_auth
def audit_logs():
    event_filter = request.args.get("event", "")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))
    result = _store().get_audit_logs(event_filter, page, page_size)
    return jsonify({"code": 0, "data": result})


# ── 告警 ─────────────────────────────────────────────────────────────────────

@app.route("/api/alert-thresholds")
@_require_auth
def get_alert_thresholds():
    return jsonify({"code": 0, "data": _store().get_alert_thresholds()})


@app.route("/api/alert-thresholds", methods=["POST"])
@_require_auth
def save_alert_thresholds():
    body = request.json or {}
    thresholds = {
        "cpu": max(1, min(100, int(body.get("cpu", 80)))),
        "mem": max(1, min(100, int(body.get("mem", 85)))),
        "disk": max(1, min(100, int(body.get("disk", 90)))),
    }
    _store().save_alert_thresholds(thresholds)
    return jsonify({"code": 0, "msg": "告警阈值已保存"})


@app.route("/api/alerts")
@_require_auth
def get_alerts():
    store = _store()
    thresholds = store.get_alert_thresholds()
    data = store.get_alerts(thresholds)
    return jsonify({"code": 0, "data": data})


# ── 邮件发送 ──────────────────────────────────────────────────────────────────

def _build_expiry_email_html(resources: list[dict], recipient_email: str) -> str:
    rows_html = ""
    for r in resources:
        days_left = r.get("days_left", 0)
        if days_left <= 7:
            badge_color = "#ef5350"
        elif days_left <= 15:
            badge_color = "#ff9800"
        elif days_left <= 30:
            badge_color = "#fdd835"
            badge_text_color = "#333"
        else:
            badge_color = "#66bb6a"
        badge_text_color = "#fff" if days_left > 30 else ("#fff" if days_left <= 15 else "#333")
        rows_html += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("name", "")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;text-align:center;">
            <span style="background:#1565c0;color:#fff;padding:2px 8px;border-radius:4px;font-size:12px;">{r.get("resource_type","")}</span>
          </td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;color:#546e7a;font-size:12px;">{r.get("n_ip","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("cloud","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("region","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("project","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("group","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("manager","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;">{r.get("expire_time","")}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #e8edf5;text-align:center;">
            <span style="background:{badge_color};color:{badge_text_color};padding:3px 10px;border-radius:10px;font-weight:700;font-size:13px;">{days_left} 天</span>
          </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:'Microsoft YaHei',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:30px 0;">
    <tr><td align="center">
      <table width="760" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.1);">

        <!-- Header -->
        <tr>
          <td style="background:linear-gradient(135deg,#1565c0,#0d47a1);padding:28px 36px;">
            <h1 style="margin:0;color:#fff;font-size:22px;font-weight:700;">☁ 云资源到期提醒</h1>
            <p style="margin:8px 0 0;color:#90caf9;font-size:14px;">以下云资源即将到期，请及时处理续费事宜</p>
          </td>
        </tr>

        <!-- Stats bar -->
        <tr>
          <td style="background:#e3f2fd;padding:14px 36px;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td style="color:#1565c0;font-size:14px;">
                  📋 共 <strong>{len(resources)}</strong> 条资源到期提醒 &nbsp;|&nbsp;
                  🚨 7天内到期：<strong style="color:#ef5350;">{sum(1 for r in resources if r.get("days_left",0) <= 7)}</strong> 条 &nbsp;|&nbsp;
                  ⚠️ 15天内到期：<strong style="color:#ff9800;">{sum(1 for r in resources if r.get("days_left",0) <= 15)}</strong> 条
                </td>
                <td align="right" style="color:#5a7aac;font-size:13px;">发送时间：{time.strftime('%Y-%m-%d %H:%M')}</td>
              </tr>
            </table>
          </td>
        </tr>

        <!-- Table -->
        <tr>
          <td style="padding:24px 36px 0;">
            <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border:1px solid #e8edf5;border-radius:6px;overflow:hidden;">
              <thead>
                <tr style="background:#1565c0;">
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;">资源名称</th>
                  <th style="padding:12px 14px;color:#fff;text-align:center;font-size:13px;font-weight:600;white-space:nowrap;">类型</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">内网IP</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">云账号</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">区域</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">项目</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">部门</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">负责人</th>
                  <th style="padding:12px 14px;color:#fff;text-align:left;font-size:13px;font-weight:600;white-space:nowrap;">到期时间</th>
                  <th style="padding:12px 14px;color:#fff;text-align:center;font-size:13px;font-weight:600;white-space:nowrap;">剩余天数</th>
                </tr>
              </thead>
              <tbody>
                {rows_html}
              </tbody>
            </table>
          </td>
        </tr>

        <!-- Footer -->
        <tr>
          <td style="padding:24px 36px 30px;">
            <div style="background:#fff8e1;border-left:4px solid #ffb300;padding:14px 18px;border-radius:4px;margin-bottom:16px;">
              <p style="margin:0;color:#e65100;font-size:13px;font-weight:600;">⚡ 温馨提示</p>
              <p style="margin:6px 0 0;color:#6d4c41;font-size:13px;">如需继续使用如上资源，请及时申请云资源费用并联系智算云计算部进行续费处理,避免资源到期导致服务中断。谢谢！</p>
            </div>
            <p style="margin:0;color:#9e9e9e;font-size:12px;text-align:center;">此邮件由中科闻歌云监控系统自动发送，请勿直接回复。</p>
          </td>
        </tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""


def _send_email(to_addr: str, subject: str, html_body: str, email_cfg=None):
    cfg = email_cfg or _cfg.get("email", {})
    host = cfg.get("host", "")
    port = int(cfg.get("port", 465))
    username = cfg.get("username", "")
    password = cfg.get("password", "")
    from_name = cfg.get("from_name", "云监控系统")
    use_ssl = cfg.get("ssl", True)

    if not host or not username or not password:
        raise ValueError("邮件服务器未配置（host/username/password 不能为空）")

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"{from_name} <{username}>"
    msg["To"] = to_addr
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    if use_ssl:
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(host, port, context=ctx, timeout=15) as server:
            server.login(username, password)
            server.sendmail(username, [to_addr], msg.as_string())
    else:
        with smtplib.SMTP(host, port, timeout=15) as server:
            server.starttls()
            server.login(username, password)
            server.sendmail(username, [to_addr], msg.as_string())


def _build_alert_email_html(items: dict, recipient_email: str) -> str:
    cpu_list = items.get("cpu", [])
    mem_list = items.get("mem", [])
    disk_list = items.get("disk", [])
    down_list = items.get("down", [])
    total = len(cpu_list) + len(mem_list) + len(disk_list) + len(down_list)

    def _common_cols(r: dict) -> str:
        return (
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("name","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#546e7a;font-size:12px;">{r.get("n_ip","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("cloud","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("region","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("project","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("group","")}</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;">{r.get("manager","")}</td>'
        )

    def _section(title: str, header_color: str, rows_html: str, headers: list[str]) -> str:
        ths = "".join(
            f'<th style="padding:10px 12px;color:#fff;text-align:left;font-size:12px;font-weight:600;white-space:nowrap;">{h}</th>'
            for h in headers
        )
        return f"""
        <tr><td style="padding:16px 36px 0;">
          <div style="background:{header_color};padding:10px 16px;border-radius:6px 6px 0 0;">
            <span style="color:#fff;font-size:14px;font-weight:600;">{title}</span>
          </div>
          <table width="100%" cellpadding="0" cellspacing="0"
            style="border-collapse:collapse;border:1px solid #e8edf5;border-top:none;border-radius:0 0 6px 6px;overflow:hidden;margin-bottom:8px;">
            <thead><tr style="background:{header_color}cc;">{ths}</tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
        </td></tr>"""

    sections_html = ""

    if cpu_list:
        rows = "".join(
            f"<tr>{_common_cols(r)}"
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#ef5350;font-weight:700;text-align:center;">{r.get("cpu_avg", 0):.2f}%</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#ef5350;font-weight:700;text-align:center;">{r.get("cpu_max", 0):.2f}%</td>'
            "</tr>"
            for r in cpu_list
        )
        sections_html += _section(
            f"🔴 CPU 使用率过高（{len(cpu_list)} 台）", "#c62828", rows,
            ["实例名称", "内网IP", "云账号", "区域", "项目", "部门", "负责人", "CPU均值%", "CPU峰值%"],
        )

    if mem_list:
        rows = "".join(
            f"<tr>{_common_cols(r)}"
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#ff9800;font-weight:700;text-align:center;">{r.get("mem_avg", 0):.2f}%</td>'
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#ff9800;font-weight:700;text-align:center;">{r.get("mem_max", 0):.2f}%</td>'
            "</tr>"
            for r in mem_list
        )
        sections_html += _section(
            f"🟠 内存使用率过高（{len(mem_list)} 台）", "#e65100", rows,
            ["实例名称", "内网IP", "云账号", "区域", "项目", "部门", "负责人", "内存均值%", "内存峰值%"],
        )

    if disk_list:
        rows = "".join(
            f"<tr>{_common_cols(r)}"
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;color:#ff9800;font-weight:700;text-align:center;">{r.get("disk_avg", 0):.2f}%</td>'
            "</tr>"
            for r in disk_list
        )
        sections_html += _section(
            f"🟡 磁盘使用率过高（{len(disk_list)} 台）", "#f57f17", rows,
            ["实例名称", "内网IP", "云账号", "区域", "项目", "部门", "负责人", "磁盘最大%"],
        )

    if down_list:
        rows = "".join(
            f"<tr>{_common_cols(r)}"
            f'<td style="padding:8px 12px;border-bottom:1px solid #e8edf5;text-align:center;">'
            f'<span style="background:#ef5350;color:#fff;padding:2px 8px;border-radius:4px;font-size:12px;">{r.get("status","DOWN")}</span>'
            f"</td></tr>"
            for r in down_list
        )
        sections_html += _section(
            f"⛔ 实例离线（{len(down_list)} 台）", "#424242", rows,
            ["实例名称", "内网IP", "云账号", "区域", "项目", "部门", "负责人", "状态"],
        )

    return f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:'Microsoft YaHei',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:30px 0;">
    <tr><td align="center">
      <table width="800" cellpadding="0" cellspacing="0"
        style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.1);">

        <tr>
          <td style="background:linear-gradient(135deg,#b71c1c,#c62828);padding:28px 36px;">
            <h1 style="margin:0;color:#fff;font-size:22px;font-weight:700;">⚠ 云资源告警提醒</h1>
            <p style="margin:8px 0 0;color:#ffcdd2;font-size:14px;">以下云资源存在告警，请及时处理</p>
          </td>
        </tr>

        <tr>
          <td style="background:#fce4ec;padding:14px 36px;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td style="font-size:14px;">
                  {'<span style="color:#c62828;margin-right:16px;">🔴 CPU告警：<strong>' + str(len(cpu_list)) + '</strong> 台</span>' if cpu_list else ''}
                  {'<span style="color:#e65100;margin-right:16px;">🟠 内存告警：<strong>' + str(len(mem_list)) + '</strong> 台</span>' if mem_list else ''}
                  {'<span style="color:#f57f17;margin-right:16px;">🟡 磁盘告警：<strong>' + str(len(disk_list)) + '</strong> 台</span>' if disk_list else ''}
                  {'<span style="color:#424242;margin-right:16px;">⛔ 实例离线：<strong>' + str(len(down_list)) + '</strong> 台</span>' if down_list else ''}
                  &nbsp;|&nbsp; 共 <strong>{total}</strong> 条
                </td>
                <td align="right" style="color:#5a7aac;font-size:13px;">发送时间：{time.strftime('%Y-%m-%d %H:%M')}</td>
              </tr>
            </table>
          </td>
        </tr>

        {sections_html}

        <tr>
          <td style="padding:20px 36px 30px;">
            <div style="background:#fff8e1;border-left:4px solid #ffb300;padding:14px 18px;border-radius:4px;margin-bottom:16px;">
              <p style="margin:0;color:#e65100;font-size:13px;font-weight:600;">⚡ 温馨提示</p>
              <p style="margin:6px 0 0;color:#6d4c41;font-size:13px;">请相关负责人尽快排查上述告警，避免影响业务正常运行。</p>
            </div>
            <p style="margin:0;color:#9e9e9e;font-size:12px;text-align:center;">此邮件由中科闻歌云监控系统自动发送，请勿直接回复。</p>
          </td>
        </tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""


@app.route("/api/alert/send-email", methods=["POST"])
@_require_action("send_email")
def alert_send_email():
    body = request.json or {}
    alerts = body.get("alerts", {})
    extra_email_raw = (body.get("extra_email") or "").strip()
    # 支持逗号分隔多个邮箱地址
    extra_emails = [e.strip() for e in extra_email_raw.split(",") if e.strip()]
    # mode: "replace" 仅发给 extra_emails；"append" 在项目邮箱基础上额外抄送
    extra_mode = body.get("extra_email_mode", "append")

    cpu_list = alerts.get("cpu", [])
    mem_list = alerts.get("mem", [])
    disk_list = alerts.get("disk", [])
    down_list = alerts.get("down", [])
    all_items = (
        [(r, "cpu") for r in cpu_list]
        + [(r, "mem") for r in mem_list]
        + [(r, "disk") for r in disk_list]
        + [(r, "down") for r in down_list]
    )
    if not all_items:
        return jsonify({"code": 400, "msg": "当前无告警数据"}), 400

    all_alerts_by_type = {"cpu": cpu_list, "mem": mem_list, "disk": disk_list, "down": down_list}

    # ── 替换模式：忽略项目邮箱，只发给 extra_emails ─────────────────────────
    if extra_emails and extra_mode == "replace":
        grouped: dict[str, dict[str, list]] = {e: all_alerts_by_type for e in extra_emails}
        no_email_count = 0
    else:
        # ── 按项目邮箱分组 ───────────────────────────────────────────────────
        store = _store()
        project_email_map: dict[str, str] = {}
        for p in store.get_all_projects():
            email_val = (p.get("notify_email") or "").strip()
            if email_val:
                project_email_map[p.get("project", "")] = email_val

        grouped = {}
        no_email_count = 0
        for item, atype in all_items:
            proj = item.get("project", "")
            email_addr = project_email_map.get(proj, "")
            if not email_addr:
                no_email_count += 1
                continue
            if email_addr not in grouped:
                grouped[email_addr] = {"cpu": [], "mem": [], "disk": [], "down": []}
            grouped[email_addr][atype].append(item)

        # 追加模式：在项目邮箱基础上再发一份给 extra_emails（包含所有告警）
        if extra_emails and extra_mode == "append":
            for e in extra_emails:
                grouped[e] = all_alerts_by_type

        if not grouped:
            return jsonify({"code": 400, "msg": "所有告警对应的项目均未配置通知邮箱，请在项目维护页面配置通知邮箱"}), 400

    email_cfg_snapshot = copy.deepcopy(_cfg.get("email", {}))

    def _send_all():
        for addr, type_items in grouped.items():
            try:
                cnt = sum(len(v) for v in type_items.values())
                subject = f"【云监控】{cnt} 条云资源告警提醒"
                html = _build_alert_email_html(type_items, addr)
                _send_email(addr, subject, html, email_cfg=email_cfg_snapshot)
                print(f"[邮件-告警] 发送成功 → {addr}（{cnt} 条）")
            except Exception as e:
                print(f"[邮件-告警] 发送失败 → {addr}: {e}")

    threading.Thread(target=_send_all, daemon=True).start()

    total_sent = len(all_items) - no_email_count
    addr_list = list(grouped.keys())
    preview = "、".join(addr_list[:3]) + ("等" if len(addr_list) > 3 else "")
    skip_msg = f"；{no_email_count} 条告警未配置项目邮箱" if no_email_count else ""
    return jsonify({
        "code": 0,
        "msg": f"已向 {len(grouped)} 个收件人（{preview}）发送共 {total_sent} 条告警提醒{skip_msg}，邮件在后台发送中",
        "sent": len(grouped),
    })


@app.route("/api/expiry/email-override", methods=["PATCH"])
@_require_auth
def expiry_email_override():
    body = request.json or {}
    cloud = (body.get("cloud") or "").strip()
    resource_id = (body.get("resource_id") or "").strip()
    email = (body.get("email") or "").strip()
    if not cloud or not resource_id:
        return jsonify({"code": 400, "msg": "cloud 和 resource_id 不能为空"}), 400
    ok = _store().save_expiry_email_override(cloud, resource_id, email)
    return jsonify({"code": 0 if ok else 404, "msg": "已保存" if ok else "资源不存在"})


@app.route("/api/expiry/send-email", methods=["POST"])
@_require_action("send_email")
def expiry_send_email():
    body = request.json or {}
    resources = body.get("resources", [])
    if not resources:
        return jsonify({"code": 400, "msg": "未选择任何资源"}), 400

    store = _store()
    # 项目级邮箱兜底（从 projects 表）
    project_email_map: dict[str, str] = {}
    for p in store.get_all_projects():
        email_val = (p.get("notify_email") or "").strip()
        if email_val:
            project_email_map[p.get("project", "")] = email_val

    # 按收件人分组：行级 notify_email 优先 > 项目级
    grouped: dict[str, list[dict]] = {}
    no_email = []
    for r in resources:
        proj = r.get("project", "")
        row_email = (r.get("notify_email") or "").strip()
        email_addr = row_email or project_email_map.get(proj, "")
        if not email_addr:
            no_email.append(r.get("name", proj))
            continue
        grouped.setdefault(email_addr, []).append(r)

    if not grouped:
        msg = "所选资源对应的项目均未配置通知邮箱"
        if no_email:
            msg += f"，未配置邮箱的项目资源：{', '.join(no_email[:5])}"
        return jsonify({"code": 400, "msg": msg}), 400

    # 快照邮件配置，避免后台线程持有 request 上下文
    email_cfg_snapshot = copy.deepcopy(_cfg.get("email", {}))

    def _send_all():
        for addr, items in grouped.items():
            try:
                subject = f"【云监控】{len(items)} 条云资源即将到期提醒"
                html = _build_expiry_email_html(items, addr)
                _send_email(addr, subject, html, email_cfg=email_cfg_snapshot)
                print(f"[邮件] 发送成功 → {addr}（{len(items)} 条）")
            except Exception as e:
                print(f"[邮件] 发送失败 → {addr}: {e}")

    threading.Thread(target=_send_all, daemon=True).start()

    total = sum(len(v) for v in grouped.items())
    addr_list = list(grouped.keys())
    preview = "、".join(addr_list[:3]) + ("等" if len(addr_list) > 3 else "")
    skip_msg = f"；{len(no_email)} 条资源未配置邮箱" if no_email else ""
    return jsonify({
        "code": 0,
        "msg": f"已向 {len(grouped)} 个收件人（{preview}）发送共 {len(resources) - len(no_email)} 条提醒{skip_msg}，邮件在后台发送中",
        "sent": len(grouped),
    })


# ── 费用统计 ──────────────────────────────────────────────────────────────────

_PROVIDER_DISPLAY = {
    "huawei":     "华为云",
    "tencent":    "腾讯云",
    "aliyun":     "阿里云",
    "volcengine": "火山云",
}
_COST_IMPLEMENTED = {"huawei", "volcengine", "aliyun", "tencent"}


@app.route("/api/cost/providers", methods=["GET"])
@_require_auth
def cost_providers():
    """返回配置中各云厂商信息及其子账号列表"""
    seen: dict[str, dict] = {}
    for cloud in _cfg.get("clouds", []):
        p = cloud.get("provider", "")
        if not p:
            continue
        if p not in seen:
            seen[p] = {
                "provider": p,
                "display_name": _PROVIDER_DISPLAY.get(p, p),
                "implemented": p in _COST_IMPLEMENTED,
                "clouds": [],
            }
        seen[p]["clouds"].append(cloud["name"])
    return jsonify({"code": 0, "data": list(seen.values())})


@app.route("/api/cost/<provider_type>", methods=["GET"])
@_require_auth
def cost_data(provider_type: str):
    from collections import defaultdict
    from datetime import datetime as _dt

    cloud  = request.args.get("cloud", "")
    period = request.args.get("period", "month")
    now    = _dt.now()
    year   = int(request.args.get("year",    now.year))
    month  = int(request.args.get("month",   now.month))
    quarter= int(request.args.get("quarter", (now.month - 1) // 3 + 1))

    # 确定查询月份和对比月份
    if period == "month":
        curr_months = [month]
        curr_year   = year
        prev_m      = month - 1 if month > 1 else 12
        prev_year   = year if month > 1 else year - 1
        cmp_months  = [prev_m]
        cmp_year    = prev_year
    elif period == "quarter":
        sm           = (quarter - 1) * 3 + 1
        curr_months  = [sm, sm + 1, sm + 2]
        curr_year    = year
        pq           = quarter - 1 if quarter > 1 else 4
        prev_year    = year if quarter > 1 else year - 1
        psm          = (pq - 1) * 3 + 1
        cmp_months   = [psm, psm + 1, psm + 2]
        cmp_year     = prev_year
    else:  # year
        curr_months = list(range(1, 13))
        curr_year   = year
        cmp_months  = list(range(1, 13))
        cmp_year    = year - 1

    store        = _store()
    curr_data    = store.get_cost_data(cloud, curr_year, curr_months)
    cmp_data     = store.get_cost_data(cloud, cmp_year,  cmp_months)

    # 按月趋势
    trend = []
    for m in curr_months:
        total = sum(r["amount"] for r in curr_data if r["month"] == m)
        trend.append({"month": f"{curr_year}-{m:02d}", "amount": round(total, 2)})

    cmp_trend = []
    for m in cmp_months:
        total = sum(r["amount"] for r in cmp_data if r["month"] == m)
        cmp_trend.append({"month": f"{cmp_year}-{m:02d}", "amount": round(total, 2)})

    # 当期：项目 × 服务类别汇总
    proj_cats: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in curr_data:
        proj_cats[r["project"]][r["service_category"]] += r["amount"]

    # 对比：项目总费用
    cmp_proj: dict[str, float] = defaultdict(float)
    for r in cmp_data:
        cmp_proj[r["project"]] += r["amount"]

    # 服务类别汇总
    cat_totals: dict[str, float] = defaultdict(float)
    for r in curr_data:
        cat_totals[r["service_category"]] += r["amount"]

    grand_total     = sum(cat_totals.values())
    grand_total_cmp = sum(cmp_proj.values())
    vs_total = round((grand_total - grand_total_cmp) / grand_total_cmp * 100, 1) \
               if grand_total_cmp > 0 else None

    # 项目明细列表（按合计降序）
    projects = []
    for proj, cats in proj_cats.items():
        total      = sum(cats.values())
        cmp_total  = cmp_proj.get(proj, 0)
        vs = round((total - cmp_total) / cmp_total * 100, 1) if cmp_total > 0 else None
        projects.append({
            "project":  proj,
            "compute":  round(cats.get("compute",  0), 2),
            "storage":  round(cats.get("storage",  0), 2),
            "network":  round(cats.get("network",  0), 2),
            "database": round(cats.get("database", 0), 2),
            "other":    round(cats.get("other",    0), 2),
            "total":    round(total, 2),
            "vs_last":  vs,
        })
    projects.sort(key=lambda x: x["total"], reverse=True)

    return jsonify({"code": 0, "data": {
        "summary": {
            "total":    round(grand_total, 2),
            "compute":  round(cat_totals.get("compute",  0), 2),
            "storage":  round(cat_totals.get("storage",  0), 2),
            "network":  round(cat_totals.get("network",  0), 2),
            "database": round(cat_totals.get("database", 0), 2),
            "other":    round(cat_totals.get("other",    0), 2),
            "vs_last":  vs_total,
        },
        "trend":       trend,
        "cmp_trend":   cmp_trend,
        "projects":    projects,
    }})


# ── 用户管理 ──────────────────────────────────────────────────────────────────

@app.route("/api/users", methods=["GET"])
@_require_admin
def list_users():
    users = _store().list_users()
    return jsonify({"code": 0, "data": users})


@app.route("/api/users", methods=["POST"])
@_require_admin
def create_user():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    email = (data.get("email") or "").strip()
    group_id = data.get("group_id")

    if not username:
        return jsonify({"code": 400, "msg": "用户名不能为空"}), 400
    if not email:
        return jsonify({"code": 400, "msg": "邮箱不能为空"}), 400
    if _store().get_user_by_username(username):
        return jsonify({"code": 400, "msg": "用户名已存在"}), 400

    raw_password = _gen_random_password()
    password_hash = _hash_password(raw_password)

    user_id = _store().create_user({
        "username": username,
        "email": email,
        "password_hash": password_hash,
        "group_id": group_id,
    })

    # 发送欢迎邮件
    system_url = _cfg.get("system_url", "")
    group_info = ""
    if group_id:
        group = _store().get_permission_group(group_id)
        if group:
            group_info = group["name"]
    try:
        html = _build_welcome_email_html(username, raw_password, system_url, group_info=group_info)
        _send_email(email, "【云监控系统】账号创建通知", html)
    except Exception as e:
        print(f"[用户创建] 欢迎邮件发送失败: {e}")

    return jsonify({"code": 0, "data": {"_id": user_id, "username": username,
                                         "temp_password": raw_password}})


@app.route("/api/users/<user_id>", methods=["PUT"])
@_require_admin
def update_user(user_id):
    data = request.json or {}
    ok = _store().update_user(user_id, data)
    return jsonify({"code": 0 if ok else 404})


@app.route("/api/users/<user_id>", methods=["DELETE"])
@_require_admin
def delete_user(user_id):
    target = _store().get_user_by_username("")  # just a type check
    # 不允许删除自己
    store = _store()
    docs = store.list_users()
    target_doc = next((u for u in docs if u["_id"] == user_id), None)
    if target_doc and target_doc.get("username") == g.username:
        return jsonify({"code": 400, "msg": "不能删除当前登录账号"}), 400
    ok = store.delete_user(user_id)
    return jsonify({"code": 0 if ok else 404})


@app.route("/api/users/<user_id>/reset-password", methods=["POST"])
@_require_admin
def reset_user_password(user_id):
    store = _store()
    docs = store.list_users()
    target = next((u for u in docs if u["_id"] == user_id), None)
    if not target:
        return jsonify({"code": 404, "msg": "用户不存在"}), 404

    raw_password = _gen_random_password()
    store.update_user(user_id, {"password_hash": _hash_password(raw_password)})

    email = target.get("email", "")
    if email:
        system_url = _cfg.get("system_url", "")
        try:
            html = _build_welcome_email_html(target["username"], raw_password, system_url,
                                              is_reset=True)
            _send_email(email, "【云监控系统】密码重置通知", html)
        except Exception as e:
            print(f"[密码重置] 邮件发送失败: {e}")

    return jsonify({"code": 0, "msg": "密码已重置并发送邮件", "temp_password": raw_password})


@app.route("/api/profile/password", methods=["PUT"])
@_require_auth
def change_own_password():
    data = request.json or {}
    old_pw = data.get("old_password", "")
    new_pw = data.get("new_password", "")
    if not new_pw or len(new_pw) < 6:
        return jsonify({"code": 400, "msg": "新密码不能少于6位"}), 400

    store = _store()
    user = store.get_user_by_username(g.username)
    if not user or not _check_password(old_pw, user.get("password_hash", "")):
        return jsonify({"code": 403, "msg": "当前密码错误"}), 403

    store.update_user(user["_id"], {"password_hash": _hash_password(new_pw)})
    return jsonify({"code": 0, "msg": "密码已修改"})


# ── 权限组管理 ─────────────────────────────────────────────────────────────────

@app.route("/api/permission-groups", methods=["GET"])
@_require_admin
def list_permission_groups():
    groups = _store().list_permission_groups()
    return jsonify({"code": 0, "data": groups})


@app.route("/api/permission-groups", methods=["POST"])
@_require_admin
def create_permission_group():
    data = request.json or {}
    if not data.get("name"):
        return jsonify({"code": 400, "msg": "权限组名称不能为空"}), 400
    gid = _store().create_permission_group(data)
    return jsonify({"code": 0, "data": {"_id": gid}})


@app.route("/api/permission-groups/<group_id>", methods=["PUT"])
@_require_admin
def update_permission_group(group_id):
    data = request.json or {}
    ok = _store().update_permission_group(group_id, data)
    return jsonify({"code": 0 if ok else 404})


@app.route("/api/permission-groups/<group_id>", methods=["DELETE"])
@_require_admin
def delete_permission_group(group_id):
    ok = _store().delete_permission_group(group_id)
    if not ok:
        return jsonify({"code": 400, "msg": "删除失败，可能有用户正在使用此权限组"}), 400
    return jsonify({"code": 0})


def _build_welcome_email_html(username: str, password: str, system_url: str,
                               is_reset: bool = False, group_info: str = "") -> str:
    action = "密码已重置" if is_reset else "账号已创建"
    title = "账号密码重置通知" if is_reset else "账号创建通知"
    desc = "您的账号密码已被管理员重置，新密码如下：" if is_reset else "管理员已为您创建云监控系统账号，初始密码如下："
    label_style = (
        "padding:10px 16px;color:#546e7a;font-size:14px;"
        "border-bottom:1px solid #f0f0f0;width:100px;"
        "vertical-align:middle;line-height:1.5;"
    )
    value_style = (
        "padding:10px 16px;font-size:14px;border-bottom:1px solid #f0f0f0;"
        "vertical-align:middle;line-height:1.5;"
    )

    def info_row(label: str, value: str, extra_value_style: str = "") -> str:
        return (
            f'<tr>'
            f'<td style="{label_style}">{label}</td>'
            f'<td style="{value_style}{extra_value_style}">{value}</td>'
            f'</tr>'
        )

    url_section = info_row(
        "系统地址",
        f'<a href="{system_url}" style="color:#1565c0;text-decoration:none;">{system_url}</a>',
    ) if system_url else ""
    group_section = info_row("权限组", group_info, "font-weight:600;") if group_info else ""

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:'Microsoft YaHei',Arial,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:30px 0;">
    <tr><td align="center">
      <table width="520" cellpadding="0" cellspacing="0"
             style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.1);">
        <tr>
          <td style="background:linear-gradient(135deg,#1565c0,#0d47a1);padding:28px 36px;">
            <h1 style="margin:0;color:#fff;font-size:20px;">☁ 云监控系统 — {title}</h1>
          </td>
        </tr>
        <tr>
          <td style="padding:30px 36px;">
            <p style="color:#333;font-size:15px;margin:0 0 20px;">{desc}</p>
            <table cellpadding="0" cellspacing="0"
                   style="border:1px solid #e0e0e0;border-radius:6px;width:100%;border-collapse:collapse;">
              <tr style="background:#f5f5f5;">
                <td style="padding:10px 16px;font-weight:600;color:#333;font-size:13px;
                           border-bottom:1px solid #e0e0e0;width:100px;">字段</td>
                <td style="padding:10px 16px;font-weight:600;color:#333;font-size:13px;
                           border-bottom:1px solid #e0e0e0;">内容</td>
              </tr>
              {info_row("用户名", username, "font-weight:600;")}
              {info_row("初始密码", password, "font-weight:600;color:#1565c0;font-family:monospace;")}
              {url_section}
              {group_section}
            </table>
            <div style="background:#fff8e1;border-left:4px solid #ffb300;padding:12px 16px;
                        border-radius:4px;margin-top:20px;">
              <p style="margin:0;color:#e65100;font-size:13px;">
                ⚡ 请登录后及时修改密码，初始密码仅作临时使用。
              </p>
            </div>
          </td>
        </tr>
        <tr>
          <td style="padding:16px 36px;background:#f5f5f5;text-align:center;">
            <p style="margin:0;color:#9e9e9e;font-size:12px;">此邮件由中科闻歌云监控系统自动发送，请勿直接回复。</p>
          </td>
        </tr>
      </table>
    </td></tr>
  </table>
</body></html>"""


# ── 运维平台链接 ──────────────────────────────────────────────────────────────

@app.route("/api/ops-links")
@_require_auth
def list_ops_links():
    links = _store().list_ops_links()
    if g.role == "admin":
        return jsonify({"code": 0, "data": links})
    # 普通用户：只返回有权限的链接
    allowed = set(g.menus)
    if "*" in allowed:
        return jsonify({"code": 0, "data": links})
    visible = [lk for lk in links if f"ops_link:{lk['_id']}" in allowed]
    return jsonify({"code": 0, "data": visible})


@app.route("/api/ops-links", methods=["POST"])
@_require_admin
def create_ops_link():
    data = request.json or {}
    if not data.get("name") or not data.get("url"):
        return jsonify({"code": 400, "msg": "name 和 url 不能为空"}), 400
    link_id = _store().create_ops_link(data)
    return jsonify({"code": 0, "id": link_id})


@app.route("/api/ops-links/<link_id>", methods=["PUT"])
@_require_admin
def update_ops_link(link_id):
    data = request.json or {}
    ok = _store().update_ops_link(link_id, data)
    if not ok:
        return jsonify({"code": 404, "msg": "链接不存在"}), 404
    return jsonify({"code": 0})


@app.route("/api/ops-links/<link_id>", methods=["DELETE"])
@_require_admin
def delete_ops_link(link_id):
    ok = _store().delete_ops_link(link_id)
    if not ok:
        return jsonify({"code": 404, "msg": "链接不存在"}), 404
    return jsonify({"code": 0})


# ── 前端静态资源 ──────────────────────────────────────────────────────────────

def _frontend_dist() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "../frontend/dist"))


def _build_version() -> str:
    """取 index.html 的 mtime 后8位作为构建版本号，每次重建后自动变更"""
    try:
        mtime = os.path.getmtime(os.path.join(_frontend_dist(), "index.html"))
        return str(int(mtime))[-8:]
    except Exception:
        return "0"


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_frontend(path):
    dist = _frontend_dist()

    # 静态资源：直接服务，带 hash 的长期缓存
    if path and os.path.exists(os.path.join(dist, path)):
        resp = send_from_directory(dist, path)
        if any(path.endswith(ext) for ext in (".js", ".css", ".woff2", ".woff", ".ttf", ".png", ".svg", ".ico")):
            resp.headers["Cache-Control"] = "public, max-age=31536000, immutable"
        else:
            resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        return resp

    # SPA 路由：若不携带 _v 参数则 302 重定向加上版本号
    # 浏览器从未见过带版本号的 URL，强制绕过 disk cache 拿新 index.html
    v = _build_version()
    if request.args.get("_v") != v:
        target = f"/{path}?_v={v}" if path else f"/?_v={v}"
        resp = redirect(target, 302)
        resp.headers["Cache-Control"] = "no-store"
        return resp

    if os.path.exists(os.path.join(dist, "index.html")):
        resp = send_from_directory(dist, "index.html")
        resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        resp.headers["Pragma"] = "no-cache"
        resp.headers["Expires"] = "0"
        return resp

    return jsonify({"msg": "前端未构建，请执行 cd frontend && npm install && npm run build"}), 404
